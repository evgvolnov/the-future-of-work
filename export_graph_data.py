from __future__ import annotations

import json
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent
SOURCE = ROOT / "taxonomy-export.xlsx"
TARGET = ROOT / "graph-data.js"

EXCLUDED_STATUSES = {"merged", "rejected", "blocked"}
EXCLUDED_RELATION_STATUSES = {"rejected", "blocked"}
FALLBACK_COLORS = [
    "#2F80ED",
    "#27AE60",
    "#F2994A",
    "#9B51E0",
    "#EB5757",
    "#56CCF2",
    "#6FCF97",
    "#BB6BD9",
    "#1B998B",
    "#E84855",
    "#2D3047",
    "#3A86FF",
    "#FF006E",
    "#FFBE0B",
    "#00A6A6",
    "#A3E635",
    "#F97316",
    "#14B8A6",
    "#E879F9",
    "#F43F5E",
]


def scalar(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def rows(ws):
    raw = list(ws.iter_rows(values_only=True))
    headers = [str(cell or "").strip() for cell in raw[0]]
    for row in raw[1:]:
        item = {headers[i]: scalar(row[i] if i < len(row) else "") for i in range(len(headers))}
        if any(item.values()):
            yield item


def as_number(value, default=0):
    if value in ("", None):
        return default
    try:
        return float(str(value).replace(",", "."))
    except ValueError:
        return default


def file_url(path):
    text = str(path or "")
    if text.startswith("http://") or text.startswith("https://"):
        return text
    return ""


def source_type(row):
    value = str(row.get("source_type", "") or "").strip().lower()
    if value in {"pdf", "article"}:
        return value
    url = str(row.get("pdf_link_or_file", "") or "")
    if url.startswith("http://") or url.startswith("https://"):
        return "article"
    return "pdf"


def main():
    wb = load_workbook(SOURCE, read_only=True, data_only=True)

    reports = {}
    for row in rows(wb["Reports"]):
        report_id = str(row["report_id"]).strip()
        if not report_id:
            continue
        reports[report_id] = {
            "title": str(row["title"] or ""),
            "publisher": str(row["publisher"] or ""),
            "year": str(row["year"] or ""),
            "url": file_url(row["pdf_link_or_file"]),
            "sourceType": source_type(row),
        }

    themes = {}
    for row in rows(wb["Themes"]):
        name = str(row["theme_name"] or "").strip()
        if name:
            themes[name] = str(row["ui_color"] or "").strip() or FALLBACK_COLORS[len(themes) % len(FALLBACK_COLORS)]

    nodes = []
    for row in rows(wb["Concepts"]):
        status = str(row["status"] or "").strip().lower()
        concept_id = str(row["concept_id"] or "").strip()
        if not concept_id or status in EXCLUDED_STATUSES:
            continue
        theme = str(row["primary_theme"] or "Uncategorized").strip()
        if theme not in themes:
            themes[theme] = FALLBACK_COLORS[len(themes) % len(FALLBACK_COLORS)]
        nodes.append(
            {
                "id": concept_id,
                "label": str(row["canonical_name"] or concept_id),
                "type": str(row["type"] or "Concept"),
                "level": str(row["level"] or "tactical").lower(),
                "theme": theme,
                "importance": int(as_number(row["importance_1_5"], 3)),
                "confidence": round(as_number(row["confidence_0_1"], 0), 2),
                "status": str(row["status"] or ""),
                "description": str(row["description"] or ""),
                "descriptionRu": str(row.get("description_ru", "") or ""),
                "aiImpact": str(row.get("ai_impact", "") or ""),
                "source": str(row["created_from_report_id"] or ""),
            }
        )

    live_ids = {node["id"] for node in nodes}

    links = []
    for row in rows(wb["Relations"]):
        status = str(row["status"] or "").strip().lower()
        source = str(row["source_concept_id"] or "").strip()
        target = str(row["target_concept_id"] or "").strip()
        if status in EXCLUDED_RELATION_STATUSES or source == target:
            continue
        if source not in live_ids or target not in live_ids:
            continue
        links.append(
            [
                source,
                target,
                str(row["relation_type"] or "related_to"),
                int(as_number(row["strength_1_5"], 3)),
            ]
        )

    evidence = []
    for row in rows(wb["Evidence"]):
        support = str(row["supports_id"] or "").strip()
        report_id = str(row["report_id"] or "").strip()
        if support not in live_ids:
            continue
        report = reports.get(report_id, {})
        start = row["page_start"]
        end = row["page_end"]
        pages = str(start) if start == end or not end else f"{start}-{end}"
        evidence.append(
            {
                "supports": support,
                "reportId": report_id,
                "report": report.get("title", report_id),
                "pages": pages,
                "excerpt": str(row["excerpt"] or ""),
                "excerptRu": str(row.get("excerpt_ru", "") or ""),
            }
        )

    aliases = {}
    aliases_ru = {}
    for row in rows(wb["Aliases"]):
        alias = str(row.get("alias", "") or row.get("alias_or_term", "") or "").strip()
        alias_ru = str(row.get("alias_ru", "") or "").strip()
        concept_id = str(row.get("canonical_concept_id", "") or row.get("concept_id", "") or "").strip()
        if alias and concept_id in live_ids:
            aliases.setdefault(concept_id, []).append(alias)
            if alias_ru:
                aliases_ru.setdefault(concept_id, []).append(alias_ru)

    payload = {
        "generatedAt": datetime.now().isoformat(timespec="seconds"),
        "reports": reports,
        "themes": themes,
        "nodes": nodes,
        "links": links,
        "evidence": evidence,
        "aliases": aliases,
        "aliasesRu": aliases_ru,
    }

    text = "window.GRAPH_DATA = "
    text += json.dumps(payload, ensure_ascii=False, indent=2)
    text += ";\n"
    TARGET.write_text(text, encoding="utf-8")
    print(
        f"Generated {TARGET.name}: {len(nodes)} nodes, {len(links)} links, "
        f"{len(evidence)} evidence rows, {len(reports)} reports."
    )


if __name__ == "__main__":
    main()
